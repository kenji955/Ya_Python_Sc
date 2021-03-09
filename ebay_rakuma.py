from collections import OrderedDict
import os
import eel
import pandas as pd
import re
import requests
import time
from googletrans import Translator
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException, ElementNotInteractableException
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from numpy import nan
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import subprocess
from bs4 import BeautifulSoup
from selenium.webdriver.support.select import Select

# 一覧ページから取得する情報
items_id = []
items_condition = []
items_name_ja = []
items_name_result = []
items_condition_en = []
items_price = []
items_image_url = []
postage_price_list = []


# 抽出対象で変わる情報
# 商品一覧と出品リストそれぞれの抽出時に必要な要素を指定する情報を定義する。
# URL ?b=に変更
page_url_after = '&b='
# 一覧ページの読み込み待ち対象
until_element_top = 'Products__items'
# リストのタグ
items_list_tagName = 'li'
# リストのクラス名
items_list_className = 'Product'
# 価格情報
price_info_className = 'Price--buynow'
# 価格情報の文字列
price_text = '即決'
# 送料情報
postageDetail_element_text = 'postageDetailBuy'


def set_driver(driver_path):
    # Chromeドライバーの読み込み
    options = ChromeOptions()

    # 起動オプションの設定
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('log-level=3')
    options.add_argument('--ignore-ssl-errors')
    options.add_argument('--incognito')          # シークレットモードの設定を付与
    options.add_argument('--user-agent=Chrome/87.0.42.88')
    options.add_argument('--single-process')
    options.add_argument('--start-maximized')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--disable-web-security')
    options.add_argument('--disable-desktop-notifications')
    options.add_argument('--disable-application-cache')
    options.add_argument("--disable-extensions")
    options.add_argument('--lang=ja')

    # ChromeのWebDriverオブジェクトを作成する。
    return Chrome(driver_path, options=options)


class MyScraping():

    # 状態の判定
    # 与えられた文字列＝商品の状態によって特定の文字列を返す
    def judge_condition(self, string):
        condition = ""

        if string == "未使用":
            condition = "Condition:Mint."

        elif string == "未使用に近い":
            condition = "Condition:Near Mint."

        elif string == "目立った傷や汚れなし":
            condition = "Condition:Excellent."

        elif string == "やや傷や汚れあり":
            condition = "Condition:Very Good."

        elif string == "傷や汚れあり":
            condition = "Condition:Good."

        elif string == "中古":
            condition = "Condition : Very Good."

        else:
            condition = nan

        return condition

    # 翻訳関数
    # 与えられた日本語の文字列を英語にして返す
    def translate(self, string):
        translator = Translator()
        result = translator.translate(string, src='ja', dest='en')
        return result.text

    # 翻訳前の不要ワードの設定
    # 第一引数の文字列に対して、第二引数のリストに含まれる文言を削除する処理を行う
    def delete_word_before(self, word, delete_before_list):
        for delete_word in delete_before_list:
            word = word.replace(delete_word, "")

        return word

    # 翻訳後の不要ワードの設定
    def delete_word_after(self, word, delete_after_list, add_name_list):

        code_regex = re.compile(
            '[!"#$&*+-.:;<=>?@^_`{|}~「」〔〕“”〈〉『』＆＊・＄＃＠。、？！｀＋￥◆★◇☆⇒%％◎○■/:×,]')
        word = code_regex.sub(" ", word)

        word = re.sub(r'\[([\w\d ]+)\]', '', word)
        word = re.sub(r'\(([\w\d ]+)\)', '', word)
        word = re.sub(r'\【([\w\d ]+)\】', '', word)

        word = word.replace("[]", "")
        word = word.replace("【】", "")

        word_list = word.split(" ")

        for delete_word in delete_after_list:
            if delete_word.lower() in word_list:
                word_list.remove(delete_word.lower())
            if delete_word.capitalize() in word_list:
                word_list.remove(delete_word.capitalize())

        for add_word in add_name_list:
            if add_word.lower() in word_list:
                word_list.remove(add_word.lower())
            if add_word.capitalize() in word_list:
                word_list.remove(add_word.capitalize())

        return word_list

    # ワードの追加
    def add_word(self, word_num, word, add_word_list):

        if len(word) == word_num:
            return word

        elif len(word) > word_num:
            while len(word) + 1 >= word_num:
                word_list = word.split(' ')
                word = ' '.join(word_list[:-1])

            if len(word) == word_num:
                return word

            else:
                word = word + ' '
                for add_word in add_word_list:
                    if add_word in word:
                        continue
                    if len(word) + len(add_word) >= word_num:
                        return word
                    word = word + add_word + " "
                else:
                    return word

        elif len(word) < word_num:
            for add_word in add_word_list:
                if add_word in word:
                    continue
                if len(word) + len(add_word) >= word_num:
                    return word
                word = word + add_word + " "
            else:
                return word

    # 抽出対象による要素変更処理
    def element_set(self, output_check):
        # 抽出する画面で変わる情報
        global page_url_after
        global until_element_top
        global items_list_tagName
        global items_list_className
        global price_info_className
        global price_text
        global postageDetail_element_text
        if(output_check == '商品一覧'):
            # URL ?b=に変更
            page_url_after = '&b='
            # 一覧ページの読み込み待ち対象
            until_element_top = 'Products__items'
            # リストのタグ
            items_list_tagName = 'li'
            # リストのクラス名
            items_list_className = 'Product'
            # 価格情報
            price_info_className = 'Price--buynow'
            # 価格情報の文字列
            price_text = '即決'
            # 送料情報
            postageDetail_element_text = 'postageDetailBuy'
        else:
            print('check')
            # URL ?b=に変更
            page_url_after = '?b='
            # 一覧ページの読み込み待ち対象
            until_element_top = 'cf'
            # until_element_top = 'inner cf'
            # リストのタグ
            items_list_tagName = 'div'
            # リストのクラス名
            items_list_className = 'bd cf'
            # 価格情報
            price_info_className = 'Price--current'
            # 価格情報の文字列
            price_text = '現在'
            # 送料情報
            postageDetail_element_text = 'postageDetailCurrent'

    # 一覧ページから情報を取り出す
    # 商品一覧から取り出す情報
    # ・商品名
    # ・商品の状態（中古とか可,良いとか）
    # ・値段、送料
    def get_top_detail(self, driver, wait, main_url, num_list, output_style, Todofuken_name):
        # ファイルの読み込み
        delete_before_file = pd.read_csv(
            './' + eel.delete_word_before()(), header=None, names=['削除するワード_before'])
        delete_before_list = delete_before_file['削除するワード_before'].tolist()
        delete_after_file = pd.read_csv(
            './' + eel.delete_word_after()(), header=None, names=['削除するワード_after'])
        delete_after_list = delete_after_file['削除するワード_after'].tolist()
        add_word_file = pd.read_csv(
            './' + eel.fill_in_word()(), header=None, names=['追加するワード'])
        add_word_list = add_word_file['追加するワード'].tolist()

        # 商品名の処理
        add_name = eel.add_word()()
        add_name_len = 80 - len(add_name)
        add_name_split_dot = add_name.split(',')
        add_name_split_brank = add_name.replace(",", " ").split(" ")

        # 型番のみの出力を行うか判別するフラグ
        kataban_check_flug = eel.kataban_check()()

        page_num = 1
        # ページごと
        for i in range(int(num_list[0]), int(num_list[1]) + 1):
            eel.view_log_js("\n" + str(i) + "ページ目開始")
            if i == int(num_list[0]):
                page_url = main_url
            else:
                page_url = main_url + page_url_after + str(page_num)
            driver.get(page_url)
            wait.until(EC.visibility_of_element_located(
                (By.CLASS_NAME, until_element_top)))

            html = driver.page_source.encode('utf-8')

            soup = BeautifulSoup(html, 'lxml')
            items = soup.find_all(items_list_tagName,
                                  class_=items_list_className)
            page_num = (len(items) * i) + 1
            if(i == int(num_list[0])):
                page_url = main_url + page_url_after + str(page_num)
                print(page_url)
                driver.get(page_url)
                print(until_element_top)
                wait.until(EC.visibility_of_element_located(
                    (By.CLASS_NAME, until_element_top)))
                html = driver.page_source.encode('utf-8')
                soup = BeautifulSoup(html, 'lxml')

                r = requests.get(page_url)
                soup = BeautifulSoup(r.content, 'lxml')
                items = soup.find_all(items_list_tagName,
                                      class_=items_list_className)
            if i == 1 and output_style == '商品一覧':
                page_num -= 3

            item_counter = 0
            # アイテムごと
            for j, item in enumerate(items):
                item_id = item.findAll("a")[0].get("data-auction-id")
                print(item_id)
                page_url = 'https://page.auctions.yahoo.co.jp/jp/auction/' + item_id
                driver.set_window_size('1920', '1080')
                driver.get(page_url)
                wait.until(EC.visibility_of_element_located(
                    (By.CLASS_NAME, 'ProductImage__inner')))
                html = driver.page_source.encode('utf-8')
                soup = BeautifulSoup(html, 'lxml')

                r = requests.get(page_url)
                soup = BeautifulSoup(r.content, 'lxml')

                Price_info = soup.find_all(
                    "div", class_=price_info_className)
                if (len(Price_info) == 0):
                    eel.view_log_js(str(j + 1) + "/" +
                                    str(len(items)) + "商品目 " + price_text + "価格なし 除外")
                    continue

                # 商品のコンディションID M列
                # 商品の状態情報抽出
                item_condition_element = soup.find_all(
                    "tr", class_="ProductTable__row")
                # テキスト化
                item_condition_element_search_index = 0
                for k, element in enumerate(item_condition_element):
                    if element.findAll('th')[0].text == '状態':
                        item_condition_element_search_index = k

                try:
                    item_condition_text = re.sub(
                        r'[\n ]', '', item_condition_element[item_condition_element_search_index].findAll("a")[0].text)
                except IndexError:
                    continue
                # 除外対象か判別。対象の場合飛ばして次の商品へ移動
                if item_condition_text == '全体的に状態が悪い':
                    eel.view_log_js(str(j + 1) + "/" +
                                    str(len(items)) + "商品目 状態が悪いため除外")
                    continue
                # コンディションIDを設定
                item_condition_id = '3000'

                # 商品名 N列
                # 商品名抽出
                item_name_result = ''
                item_name_ja_element = soup.find_all(
                    "h1", class_="ProductTitle__text")
                item_name_ja = item_name_ja_element[0].text
                if kataban_check_flug:
                    # 型番抽出
                    # 英数字もしくはハイフンを抽出
                    item_name_kataban = re.findall(
                        '[a-zA-Z0-9_]+|-', item_name_ja, flags=re.IGNORECASE)
                    # 抽出した文字列を連結
                    for kataban in item_name_kataban:
                        result_kataban = result_kataban + kataban + ' '
                    # 末尾のスペースを削除し抽出完了
                    item_name_result = result_kataban.rstrip()

                else:
                    item_name_ja_delete = self.delete_word_before(
                        item_name_ja, delete_before_list)
                    item_name_en = self.translate(item_name_ja_delete)
                    item_name_en_delete_list = self.delete_word_after(
                        item_name_en, delete_after_list, add_name_split_brank)
                    item_name_en = ' '.join(
                        OrderedDict.fromkeys(item_name_en_delete_list))
                    item_name_result = item_name_en.title()[:80]

                # 商品識別ID O列
                # 取得済み

                # 商品の値段 Q列
                price_element = Price_info[0].findAll(
                    "dd", class_="Price__value")[0].text
                item_price = 0
                if '税込' in price_element:
                    price_element = re.split('[円税込]', price_element)[3]
                    item_price = re.sub(r'[, ]', '', price_element)
                else:
                    item_price = price_element.split('円')[0].replace(
                        ',', '').replace('\n', '')
                # 配送料チェック
                postage_price_free_element = soup.find_all(
                    "span", class_="Price__postageValue--free")
                # 配送料無料ではない商品をピックアップする。
                if (len(postage_price_free_element) == 0):
                    postage_price_text_element = Price_info[0].findAll(
                        "span", class_="Price__postageValue")[0].text
                    if('商品説明参照' not in postage_price_text_element):
                        # 料金はpostage_priceで取得し計算する
                        # postage_price_list.append(item_id)
                        try:
                            # driver.set_window_size('1920', '1080')
                            # driver.get(
                            #     "https://page.auctions.yahoo.co.jp/jp/auction/" + item)
                            # wait.until(EC.visibility_of_element_located(
                            #     (By.CLASS_NAME, "ProductInformation__item")))
                            driver.execute_script(
                                'document.querySelector("#js-prMdl-close").click()')
                            postageDetail_element = driver.find_element_by_id(
                                postageDetail_element_text)
                            postageDetail_element.click()

                            select_element = driver.find_element_by_xpath(
                                "//div[@class='SelectWrap']/select")
                            select_object = Select(select_element)
                            select_object.select_by_visible_text(
                                Todofuken_name)

                            postagePrice_element = driver.find_element_by_xpath(
                                "//dd[@class='BidModal__postageDetail']/div[@class='BidModal__postagePrice']")

                            postagePrice = (postagePrice_element.text).split('円')[
                                0].replace(',', '')
                            if(postagePrice == '送料未定'):
                                pass
                            elif(postagePrice == '送料未定（着払い）'):
                                item_price = int(item_price) + 3000
                            else:
                                item_price = int(item_price) + \
                                    int(postagePrice)
                        except ElementNotInteractableException:
                            continue
                        except NoSuchElementException:
                            continue

                # 商品画像(10枚まで) W列
                item_images_element = soup.find_all(
                    "div", class_="ProductImage__inner")
                limit = 10
                urls = ''
                for k, item_image in enumerate(item_images_element):
                    urls += item_image.findAll("img")[0].get("src") + "|"
                    if k + 1 == limit:
                        break
                image_urls = urls

                # 商品の状態 Y列
                item_condition_en = self.judge_condition(
                    item_condition_text.replace('\n', ''))

                # 各データを格納
                items_condition.append(item_condition_id)
                items_name_ja.append(item_name_ja)
                items_name_result.append(item_name_result)
                items_id.append(item_id)
                items_price.append(item_price)
                items_image_url.append(image_urls)
                items_condition_en.append(item_condition_en)

                eel.view_log_js(str(j + 1) + "/" + str(len(items)) + "商品目")
                item_counter += 1

            else:
                eel.view_log_js(str(i) + "ページ目終了")
                eel.view_log_js(str(item_counter) + "件抽出")

        return items_name_ja, items_image_url


def main():

    try:
        path = os.getcwd()
        driver_path = ChromeDriverManager(path=path).install()

        driver = set_driver(driver_path)
        driver.implicitly_wait(10)
        wait = WebDriverWait(driver, 10)
        myscraping = MyScraping()
        num_list = eel.page()().split(',')
        output_style = eel.output_check()()
        myscraping.element_set(output_style)
        Todofuken_name = eel.Todofuken_name()()
        name_list, items_image_url = myscraping.get_top_detail(
            driver, wait, eel.url()(), num_list, output_style, Todofuken_name)
        eel.view_log_js('\n商品一覧ページからの収集が完了しました')
        driver.quit()

        eel.view_log_js('\nファイルの作成をします')

        result_array = {
            "M": items_condition,
            "N": items_name_result,
            "O": items_id,
            "Q": items_price,
            "W": items_image_url,
            "Y": items_condition_en
        }
        result_df = pd.DataFrame(result_array)

        result_df.dropna(subset=['O', 'Y'], inplace=True)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws["M1"] = "*ConditionID"
        ws["N1"] = "Title"
        ws["O1"] = "CustomLabel"
        ws["Q1"] = "元値"
        ws["W1"] = "PicURL"
        ws["Y1"] = "ConditionDescription"

        for column_name, item_list in result_df.iteritems():
            i = 2
            for item in item_list:
                ws[column_name + str(i)] = item
                i += 1

        wb.save(eel.file_name()())

        eel.view_log_js('ファイルの作成が完了しました')

    except WebDriverException:
        import traceback
        t = traceback.format_exc()
        eel.view_log_js(t)
        eel.view_log_js("一度、driversファイルとdrivers.jsonを削除してからやり直してください")

    except Exception:
        import traceback
        t = traceback.format_exc()
        eel.view_log_js(t)
        eel.view_log_js("予期せぬエラーが発生しました。")
